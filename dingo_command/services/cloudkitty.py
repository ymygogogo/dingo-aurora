import json
import os
import platform
import shutil
from datetime import datetime
from urllib.parse import unquote

from openpyxl.reader.excel import load_workbook
from openpyxl.styles import Border, Side
from oslo_log import log
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import SimpleDocTemplate, Table

from dingo_command.common.cloudkitty_client import CloudKittyClient
from dingo_command.utils.constant import RATING_SUMMARY_TEMPLATE_FILE_DIR
from dingo_command.utils.datetime import utc_to_system_time, convert_timestamp_to_date
import pandas as pd
from openpyxl.utils.dataframe import dataframe_to_rows

LOG = log.getLogger(__name__)

# 定义边框样式
thin_border = Border(
    left=Side(border_style="thin", color="000000"),  # 左边框
    right=Side(border_style="thin", color="000000"),  # 右边框
    top=Side(border_style="thin", color="000000"),  # 上边框
    bottom=Side(border_style="thin", color="000000")  # 下边框
)

class CloudKittyService:

    def download_rating_summary_excel(self, result_file_path, query_params):
        # 模板路径
        current_template_file = os.getcwd() + RATING_SUMMARY_TEMPLATE_FILE_DIR
        # 对应类型的模板不存在
        if current_template_file is None:
            return None
        try:
            print(f"current_template_file:{current_template_file}, result_file_path:{result_file_path}")
            # 复制模板文件到临时目录
            shutil.copy2(current_template_file, result_file_path)
            # 计费汇总的文件
            self.query_data_and_create_rating_summary_excel(result_file_path, query_params)

        except Exception as e:
            import traceback
            traceback.print_exc()
            raise e

    def query_data_and_create_rating_summary_excel(self, result_file_path, query_params):
        try:
            # 1. 并行获取数据
            start_time = datetime.now()
            print(f"Data fetch time start: {start_time}")
            storage_dataFrames = CloudKittyClient().get_storage_dataframes(query_params)
            print(f"Data fetch time: {(datetime.now() - start_time).total_seconds():.2f}s, now time: {datetime.now()}")

            start_time2 = datetime.now()
            # 2. 使用向量化处理数据
            excel_data = []
            for temp in filter(None, storage_dataFrames):  # 过滤None
                resources = temp.get("resources", [])
                excel_data.append({
                    'Begin': temp.get("begin"),
                    'End': temp.get("end"),
                    'Resources': json.dumps(resources) if resources else None
                })
            print(f"handle time: {(datetime.now() - start_time2).total_seconds():.2f}s, now time: {datetime.now()}")
            start_time = datetime.now()
            # 3. 批量写入Excel
            book = load_workbook(result_file_path)
            sheet = book['ratingSummary']

            # 使用pandas加速并批量写入
            data_df = pd.DataFrame(excel_data)
            self._write_to_excel_bulk(sheet, data_df, start_row=2)

            book.save(result_file_path)
            print(f"Excel save time: {(datetime.now() - start_time).total_seconds():.2f}s, now time: {datetime.now()}")

        except Exception as e:
            import traceback
            traceback.print_exc()
            raise e

    # def _fetch_data_parallel(self, query_params):
    #     """并行获取数据"""
    #     with ThreadPoolExecutor(max_workers=4) as executor:
    #         return list(executor.map(lambda x: CloudKittyClient().get_storage_dataframes(x), [query_params]))[0]

    def _write_to_excel_bulk(self, sheet, data_df, start_row):
        """批量写入优化"""
        for r_idx, row in enumerate(dataframe_to_rows(data_df, index=False, header=False), start_row):
            for c_idx, value in enumerate(row, 1):
                sheet.cell(row=r_idx, column=c_idx, value=value).border = thin_border


    def generate_rating_summary_detail_pdf(self, result_file_pdf_path, temp_data):
        try:
            # 生成PDF文件
            self.create_rating_summary_detail_pdf_with_table(temp_data, result_file_pdf_path)

        except Exception as e:
            import traceback
            traceback.print_exc()
            raise e

    def generate_rating_summary_detail_data(self, ratingSummaryDetailList, language):
        try:
            if ratingSummaryDetailList is None:
                return None, None
            temp_data = []
            # 项目名称
            tenant_name = None
            tenant_id =  None
            # 开始时间
            begin_datatime = None
            # 结束时间
            end_datatime = None
            if ratingSummaryDetailList is not None and ratingSummaryDetailList[0] is not None:
                for ratingSummaryDetail in ratingSummaryDetailList:
                    if ratingSummaryDetail.service == "总计" or ratingSummaryDetail.service.lower() == "total":
                        tenant_name = ratingSummaryDetail.tenant_name
                        tenant_id = ratingSummaryDetail.tenant_id
                        begin_datatime = ratingSummaryDetail.start_time
                        end_datatime = ratingSummaryDetail.end_time

            # tenant_id_and_name_period = tenant_id + "_" + tenant_name
            tenant_id_and_name_period = tenant_name
            if language is None or language == "EN":
                temp_data.append({'Tenant Name': tenant_name + f"(ID: {tenant_id})"})
                temp_data.append({'Begin Time': begin_datatime})
                temp_data.append({'End Time': end_datatime})
                # tenant_id_and_name_period = tenant_id_and_name_period + "_" + switch_time_to_time(begin_datatime, end_datatime)
                tenant_id_and_name_period = tenant_id_and_name_period + "_" + convert_timestamp_to_date(unquote(begin_datatime), unquote(end_datatime))
                # Res type、Rate
                temp_data.append({'Res Type': 'Rate'})
            else:
                temp_data.append({'项目名称': tenant_name + f"(ID: {tenant_id})"})
                temp_data.append({'开始时间': begin_datatime})
                temp_data.append({'结束时间': end_datatime})
                # Res type、费率
                temp_data.append({'资源类型': '费率'})

            temp_instance_flavor_data = []
            temp_non_instance_data = []
            temp_total_rating_data = []
            for ratingSummaryDetail in ratingSummaryDetailList:
                if ratingSummaryDetail.service == "总计" or ratingSummaryDetail.service.lower() == "total":
                    if language is None or language == "EN":
                        temp_total_rating_data.append({'Total':ratingSummaryDetail.total})
                    else:
                        temp_total_rating_data.append({'总计':ratingSummaryDetail.total})
                elif ratingSummaryDetail.service == "instance":
                    temp_instance_value = f"{ratingSummaryDetail.total or '0'}(CPU:{ratingSummaryDetail.cpuTotal or '0'}, GPU:{ratingSummaryDetail.gpuTotal or '0'})"
                    temp_instance_flavor_data.append({'instance':temp_instance_value})
                    # if ratingSummaryDetail.flavor is not None:
                    #     for flavor in ratingSummaryDetail.flavor:
                    #         flavor_name_final = "instance-flavor-"
                    #         if flavor.flavor_name is None:
                    #             flavor_name_final = flavor_name_final + "None"
                    #         else:
                    #             flavor_name_final = flavor_name_final + flavor.flavor_name
                    #         temp_instance_flavor_data.append({flavor_name_final: flavor.rate})
                else:
                    temp_non_instance_data.append({ratingSummaryDetail.service:ratingSummaryDetail.total})

            # 整合所有数据
            if temp_non_instance_data is not None:
                temp_data.extend(temp_non_instance_data)
            if temp_instance_flavor_data is not None:
                temp_data.extend(temp_instance_flavor_data)
            if temp_total_rating_data is not None:
                temp_data.extend(temp_total_rating_data)
            # print(f"整合后的数据：{temp_data}")
            return temp_data, tenant_id_and_name_period
        except Exception as e:
            import traceback
            traceback.print_exc()
            raise RuntimeError(f"生成PDF数据: {e}")

    def create_rating_summary_detail_pdf_with_table(self, input_data, filename):
        try:
            # 转换数据为二维数组
            # table_data = [[list(item.keys())[0], list(item.values())[0]] for item in input_data]

            is_has_instance_flavor_prefix = self.has_instance_flavor_prefix(input_data)
            table_data = []
            instance_total_value = 0
            for i, item in enumerate(input_data):
                key = list(item.keys())[0]
                value = list(item.values())[0]

                if i < 3:  # 前三行
                    table_data.append([key, value, "", ""])  # 添加空列占位
                elif key.startswith('instance'):  # instance前缀行
                    if key == "instance":
                        instance_total_value = value
                        if not is_has_instance_flavor_prefix:
                            table_data.append(["instance", value,"", ""])
                    # else:
                        # table_data.append(["instance", key.removeprefix("instance-flavor-"), value, instance_total_value])  # 同样四列
                else:
                    table_data.append([key, value, "", ""])

            # 创建PDF文档
            doc = SimpleDocTemplate(filename, pagesize=A4)

            if platform.system() == "Windows":
                # 注册中文字体
                pdfmetrics.registerFont(TTFont('SimSun', 'simsun.ttc'))
                font_name = 'SimSun'
                print(f"system: {platform.system()}, font_name: {font_name}")
            else:
                try:
                    # 尝试注册系统可能存在的中文字体
                    pdfmetrics.registerFont(TTFont('SimSun', '/usr/share/fonts/chinese/simsun.ttc'))
                    font_name = 'SimSun'
                    print(f"system: {platform.system()}, font_name: {font_name}")
                except Exception as e:
                    font_name = 'Helvetica'
                    print(f"system: {platform.system()}, font_name: {font_name}")
                    import traceback
                    traceback.print_exc()
            print(f"registered font:{pdfmetrics.getRegisteredFontNames()}, font_name:{font_name}")

            # 定义表格样式（包含列宽设置）
            col_widths = [120, 150, 100, 100] if len(table_data[0]) > 2 else [120, 150]

            # print(f"table data:{table_data}")
            table = Table(table_data, colWidths=col_widths)
            # 基础样式
            style = [
                ('FONTNAME', (0, 0), (-1, -1), font_name),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT')
            ]

            # 前三行不显示网格线
            style.append(('GRID', (0, 3), (-1, -1), 1, colors.black))

            # 处理instance行的特殊合并
            first_instance_row = None
            fourth_instance_row = None
            for i, row in enumerate(table_data):
                if row[0] != "instance":
                    # 合并非instance行的后三列
                    style.append(('SPAN', (1, i), (3, i)))
                    # 重置instance行标记
                    first_instance_row = None
                    fourth_instance_row = None
                else:
                    if is_has_instance_flavor_prefix == False:
                        style.append(('SPAN', (1, i), (3, i)))
                         # 重置instance行标记
                        first_instance_row = None
                        fourth_instance_row = None
                    else:
                        # 处理第一列合并
                        if first_instance_row is None:
                            first_instance_row = i
                        else:
                            style.append(('SPAN', (0, first_instance_row), (0, i)))
                            table_data[i][0] = ""  # 清空后续行内容

                        # 处理第四列合并
                        if fourth_instance_row is None:
                            fourth_instance_row = i
                        else:
                            style.append(('SPAN', (3, fourth_instance_row), (3, i)))
                            table_data[i][3] = ""  # 清空后续行内容

            table.setStyle(style)

            # 构建PDF
            doc.build([table])
        except Exception as e:
            import traceback
            traceback.print_exc()
            raise RuntimeError(f"生成PDF[{filename}]失败: {e}")

    def has_instance_flavor_prefix(self, data):
        """
        检查JSON数据中是否存在以'instance-flavor-'为前缀的key
        :param data: 输入的JSON数据
        :return: 布尔值，表示是否存在符合条件的key
        """
        for item in data:
            for key in item.keys():
                if key.startswith('instance-flavor-'):
                    return True
        return False

    def edit_rating_module_config_hashmap_mappings(self, mapping_id, mapping):
        try:
            return CloudKittyClient().modify_rating_module_config_hashmap_mappings(mapping_id, mapping)
        except Exception as e:
            import traceback
            traceback.print_exc()
            raise RuntimeError(f"modify rating module config hashmap mapping fail: {e}")

    def edit_rating_module_config_hashmap_thresholdings(self, threshold_id, thresholding):
        try:
            return CloudKittyClient().modify_rating_module_config_hashmap_thresholdings(threshold_id, thresholding)
        except Exception as e:
            import traceback
            traceback.print_exc()
            raise RuntimeError(f"modify rating module config hashmap thresholding fail: {e}")

    def edit_rating_module_modules(self, module_id, modules):
        try:
            return CloudKittyClient().edit_rating_module_modules(module_id, modules)
        except Exception as e:
            import traceback
            traceback.print_exc()
            raise RuntimeError(f"modify rating modules fail: {e}")

    def get_rating_report_summary(self, filters):
        try:
            openstack_data = CloudKittyClient().get_rating_report_summary(filters)
            result = {}
            # 处理数据
            result_data = []
            for item in filter(None, openstack_data['summary']):
                if not isinstance(item, dict):
                    continue

                single_data = {}
                for key, value in item.items():
                    if key == "begin" or key == "end":
                        single_data[key] = utc_to_system_time(value)
                    else:
                        single_data[key] = value


                result_data.append(single_data)
            result['summary'] = result_data
            return result
        except Exception as e:
            import traceback
            traceback.print_exc()
            raise RuntimeError(f"get rating report summary fail: {e}")

    def get_rating_report_total(self, filters):
        try:
            return CloudKittyClient().get_rating_report_total(filters)
        except Exception as e:
            import traceback
            traceback.print_exc()
            raise RuntimeError(f"get rating report total fail: {e}")